from openpyxl import Workbook

wb = Workbook()
ws = wb.active
print(ws.max_row)
val = ws.cell(row=1, column=1).value
print(val)
ws.append(["header1", "header2"])
wb.save("test.xlsx")
