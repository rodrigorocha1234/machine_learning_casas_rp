from openpyxl import load_workbook
wb = load_workbook("test.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
