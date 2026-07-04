from openpyxl import Workbook
wb = Workbook()
ws = wb.active
if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
    for i, cabecalho in enumerate(["header1", "header2"], start=1):
        ws.cell(row=1, column=i, value=cabecalho)
ws.append(["val1", "val2"])
wb.save("test2.xlsx")

from openpyxl import load_workbook
wb = load_workbook("test2.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
