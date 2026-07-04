from typing import Dict, Union, List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.dados.arquivo import Arquivo

DadosExcel = list[dict[str, str | int]]


class ArquivoExcel(Arquivo[Workbook, DadosExcel]):
    def __init__(self, nome_pasta_amarzenamento: str, nome_arquivo: str, nome_aba: str):
        super().__init__(nome_pasta_amarzenamento, nome_arquivo)
        self.__planilha = Workbook()
        self.__nome_aba = nome_aba

    def __criar_cabecalho(self, dados: List[Dict[str, Union[str, int]]], aba: Worksheet) -> List[str]:
        """Métod para criar o cabeçalho das colunas

        Args:
            dados (Dict[str, Union[str, int]]): dados da req
            aba (worksheet): nome da aba

        Returns:
            List[str]: Lista de cabeçalhos
        """
        cabecalhos = list(dados[0].keys())
        
        if aba.max_row == 1 and aba.cell(row=1, column=1).value is None:
            for col_idx, cabecalho in enumerate(cabecalhos, start=1):
                aba.cell(row=1, column=col_idx, value=cabecalho)

        return cabecalhos

    def definir_aba(self, nome_aba: str):
        self.__nome_aba = nome_aba

    def salvar_dados(self, dados: List[Dict[str, Union[str, int]]]):
        """Método para salvar os dados da planilha

        Args:
            dados (Dict[str, Union[str, int]]): dados
        """

        if self.__nome_aba in self.__planilha.sheetnames:
            aba = self.__planilha[self.__nome_aba]
        else:
            if self.__planilha.sheetnames == ["Sheet"]:
                active_sheet = self.__planilha.active
                if active_sheet is None:
                    raise RuntimeError("Workbook sem planilha ativa.")
                aba = active_sheet
                aba.title = self.__nome_aba
            else:
                aba = self.__planilha.create_sheet(self.__nome_aba)

        cabecalhos = self.__criar_cabecalho(dados=dados, aba=aba)
        for linha in dados:
            valores = [linha[coluna] for coluna in cabecalhos]
            aba.append(valores)
        self.__planilha.save(self._caminho_arquivo)

    def atualizar_dados(self, dados: List[Dict[str, Union[str, int]]]):
        """Método para atualizar dados da planilha

        Args:
            dados (Dict[str, Union[str, int]]): dados da req
        """
        workbook = load_workbook(self._caminho_arquivo)
        if self.__nome_aba not in workbook.sheetnames:
            planilha = workbook.create_sheet(self.__nome_aba)
            cabecalhos = self.__criar_cabecalho(dados=dados, aba=planilha)
            for linha in dados:
                valores = [linha[coluna] for coluna in cabecalhos]
                planilha.append(valores)
        else:
            planilha = workbook[self.__nome_aba]
            ultima_lina = planilha.max_row + 1
            for _, valor in enumerate(dados, start=ultima_lina):
                planilha.append(list(valor.values()))

            ultima_lina = planilha.max_row

        workbook.save(self._caminho_arquivo)
        workbook.close()
